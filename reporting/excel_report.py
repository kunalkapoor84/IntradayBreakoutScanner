from datetime import datetime
from typing import List, Set
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.settings import CONFIG
from config.logging_setup import setup_logging
from models import ScoredStock, ScannerOutput, Direction
from scoring.ai_scorer import AIScorer

logger = setup_logging("excel_report")


class ExcelReportGenerator:
    def generate(self, output: ScannerOutput, filepath: str = ""):
        if not filepath:
            filepath = f"{CONFIG.output_dir}\\reports\\scanner_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = Workbook()
        self._write_full_list(wb, output)
        self._write_top_picks(wb, output)
        wb.save(filepath)
        logger.info(f"Excel saved: {filepath}")
        return filepath

    def _header_style(self):
        return {
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "font": Font(bold=True, color="FFFFFF"),
            "align": Alignment(horizontal="center"),
            "border": Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")),
        }

    def _cell_border(self):
        return Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def _write_header(self, ws, headers):
        hs = self._header_style()
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hs["fill"]; c.font = hs["font"]; c.alignment = hs["align"]; c.border = hs["border"]

    def _write_rows(self, ws, stocks, selected_symbols: Set[str], headers, show_selected: bool = True):
        tb = self._cell_border()
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        scorer = AIScorer()
        for ri, s in enumerate(stocks, 2):
            band = scorer.get_confidence_band(s.total_score)
            selected = "Yes" if s.symbol in selected_symbols else "No"
            direction = s.direction.value if hasattr(s.direction, "value") else str(s.direction)
            atr_pct = round(s.atr / max(s.cmp, 1) * 100, 2) if s.atr else 0
            extras = ""
            if s.bollinger_squeeze: extras += "BB-SQZ "
            if s.nr_detected: extras += "NR "
            if s.breakout_proximity <= 2: extras += f"BRK-{s.breakout_proximity:.0f}%"
            extras = extras.strip()
            vals = [
                ri - 1, s.symbol, s.cmp, s.entry_price, round(s.total_score, 1), band, round(s.confidence, 1),
                direction, s.stop_loss, s.target_1, s.target_2,
                round(s.atr, 2), atr_pct, round(s.volume_ratio, 2), round(s.rsi, 1),
                round(s.adx, 1), s.pcr, s.oi_change, round(s.iv, 1),
                s.sector, s.strategy, s.pattern_detected, extras,
                round(s.risk_reward, 2), s.position_size, round(s.expected_move_pct, 2),
                s.catalyst,
            ]
            if show_selected:
                vals.append(selected)
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = tb; c.alignment = Alignment(horizontal="center")
            if selected == "Yes" and show_selected:
                for ci in range(1, len(vals) + 1):
                    ws.cell(row=ri, column=ci).fill = yellow
            dc = ws.cell(row=ri, column=8)
            if "Bullish" in str(dc.value): dc.fill = green
            elif "Bearish" in str(dc.value): dc.fill = red
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

    def _write_full_list(self, wb, output):
        ws = wb.active
        ws.title = "Full List"
        headers = [
            "Rank", "Symbol", "CMP", "Entry", "Score", "Band", "Conf%", "Direction",
            "SL", "T1", "T2", "ATR", "ATR%", "Vol Ratio", "RSI",
            "ADX", "PCR", "OI Chg", "IV", "Sector", "Strategy", "Pattern",
            "Extras", "R:R", "Pos Size", "Exp Move%", "Catalyst", "Selected",
        ]
        all_sorted = sorted(output.all_scored, key=lambda s: s.total_score, reverse=True)
        selected = {s.symbol for s in output.top_stocks}
        self._write_header(ws, headers)
        self._write_rows(ws, all_sorted, selected, headers, show_selected=True)

    def _write_top_picks(self, wb, output):
        ws = wb.create_sheet("Top Picks")
        headers = [
            "Rank", "Symbol", "CMP", "Entry", "Score", "Band", "Conf%", "Direction",
            "SL", "T1", "T2", "ATR", "ATR%", "Vol Ratio", "RSI",
            "ADX", "PCR", "OI Chg", "IV", "Sector", "Strategy", "Pattern",
            "Extras", "R:R", "Pos Size", "Exp Move%", "Catalyst",
        ]
        selected = {s.symbol for s in output.top_stocks}
        self._write_header(ws, headers)
        self._write_rows(ws, output.top_stocks, selected, headers, show_selected=False)
